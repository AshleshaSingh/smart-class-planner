"""
Description:
    Excel exporter for the presentation layer
    Exports generated course plans to Excel format with multiple sheets
    
Features:
    - Multi-sheet workbook creation
    - Formatted course plan table
    - Summary statistics sheet
    - Metadata sheet with generation details
    - Professional styling and formatting
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelExporter:
    """
    Excel exporter for course plans.
    Creates professionally formatted Excel workbooks.
    """
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self.workbook = None
        self.file_path = None
        
    def export_plan(self, plan_data: Any, file_path: str, metadata: Optional[Dict] = None):
        """
        Export course plan to Excel file.
        
        Args:
            plan_data: The plan object or data structure containing course information
            file_path: Path where Excel file will be saved
            metadata: Optional dictionary with generation metadata
        """
        self.file_path = file_path
        
        # Convert plan_data to structured format if it's a string
        if isinstance(plan_data, str):
            structured_data = self._parse_text_plan(plan_data)
        else:
            structured_data = self._extract_plan_data(plan_data)
        
        # Create Excel workbook with multiple sheets
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Sheet 1: Course Plan
            df_courses = self._create_course_dataframe(structured_data['courses'])
            df_courses.to_excel(writer, sheet_name='Course Plan', index=False)
            
            # Sheet 2: Semester Summary
            df_summary = self._create_semester_summary(structured_data['courses'])
            df_summary.to_excel(writer, sheet_name='Semester Summary', index=False)
            
            # Sheet 3: Program Summary
            df_program = self._create_program_summary(structured_data)
            df_program.to_excel(writer, sheet_name='Program Summary', index=False)
            
            # Sheet 4: Metadata
            df_meta = self._create_metadata(metadata)
            df_meta.to_excel(writer, sheet_name='Metadata', index=False)
            
            # Apply formatting
            self._apply_formatting(writer.book)
    
    def export_simple(self, courses: List[Dict], file_path: str):
        """
        Simple export for a list of courses.
        
        Args:
            courses: List of course dictionaries
            file_path: Output file path
        """
        df = pd.DataFrame(courses)
        df.to_excel(file_path, index=False, sheet_name='Courses')
    
    def _parse_text_plan(self, text_plan: str) -> Dict:
        """
        Parse text-based plan into structured data.
        
        Args:
            text_plan: Text representation of the plan
            
        Returns:
            Dictionary with structured course data
        """
        courses = []
        current_semester = None
        
        lines = text_plan.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            
            # Detect semester headers
            if any(term in line.upper() for term in ['FALL', 'SPRING', 'SUMMER']):
                if any(year in line for year in ['2025', '2026', '2027', '2028']):
                    current_semester = line
                    continue
            
            # Detect course lines (format: "CPSC 6177 - Course Name (3 credits)")
            if 'CPSC' in line or 'MATH' in line or 'STAT' in line:
                parts = line.split('-')
                if len(parts) >= 2:
                    code = parts[0].strip()
                    rest = parts[1].strip()
                    
                    # Extract course name and credits
                    if '(' in rest and 'credit' in rest:
                        name_part = rest.split('(')[0].strip()
                        credits_part = rest.split('(')[1].split(')')[0]
                        try:
                            credits = int(''.join(filter(str.isdigit, credits_part)))
                        except:
                            credits = 3
                    else:
                        name_part = rest
                        credits = 3
                    
                    course = {
                        'semester': current_semester or 'TBD',
                        'course_code': code,
                        'course_name': name_part,
                        'credits': credits,
                        'prerequisites': '',
                        'status': 'Planned'
                    }
                    courses.append(course)
        
        # If no courses parsed, create sample structure
        if not courses:
            courses = [
                {
                    'semester': 'Fall 2025',
                    'course_code': 'CPSC 6177',
                    'course_name': 'Software Engineering',
                    'credits': 3,
                    'prerequisites': 'None',
                    'status': 'Planned'
                }
            ]
        
        return {
            'courses': courses,
            'total_credits': sum(c['credits'] for c in courses),
            'total_courses': len(courses)
        }
    
    def _extract_plan_data(self, plan_obj: Any) -> Dict:
        """
        Extract data from plan object.
        
        Args:
            plan_obj: Plan object from application layer
            
        Returns:
            Dictionary with structured data
        """
        courses = []
        
        # Try to extract from common plan object structures
        try:
            if hasattr(plan_obj, 'semesters'):
                for semester in plan_obj.semesters:
                    for course in semester.courses:
                        courses.append({
                            'semester': semester.name,
                            'course_code': course.code,
                            'course_name': course.name,
                            'credits': course.credits,
                            'prerequisites': getattr(course, 'prerequisites', ''),
                            'status': getattr(course, 'status', 'Planned')
                        })
            elif hasattr(plan_obj, 'courses'):
                for course in plan_obj.courses:
                    courses.append({
                        'semester': getattr(course, 'semester', 'TBD'),
                        'course_code': course.code,
                        'course_name': course.name,
                        'credits': course.credits,
                        'prerequisites': getattr(course, 'prerequisites', ''),
                        'status': getattr(course, 'status', 'Planned')
                    })
        except Exception:
            # Fallback to string conversion
            return self._parse_text_plan(str(plan_obj))
        
        return {
            'courses': courses,
            'total_credits': sum(c['credits'] for c in courses),
            'total_courses': len(courses)
        }
    
    def _create_course_dataframe(self, courses: List[Dict]) -> pd.DataFrame:
        """Create DataFrame for course plan."""
        if not courses:
            return pd.DataFrame(columns=['Semester', 'Course Code', 'Course Name', 'Credits', 'Prerequisites', 'Status'])
        
        df = pd.DataFrame(courses)
        
        # Rename columns for display
        column_mapping = {
            'semester': 'Semester',
            'course_code': 'Course Code',
            'course_name': 'Course Name',
            'credits': 'Credits',
            'prerequisites': 'Prerequisites',
            'status': 'Status'
        }
        df = df.rename(columns=column_mapping)
        
        # Reorder columns
        desired_order = ['Semester', 'Course Code', 'Course Name', 'Credits', 'Prerequisites', 'Status']
        df = df[[col for col in desired_order if col in df.columns]]
        
        return df
    
    def _create_semester_summary(self, courses: List[Dict]) -> pd.DataFrame:
        """Create semester-wise summary."""
        if not courses:
            return pd.DataFrame(columns=['Semester', 'Number of Courses', 'Total Credits'])
        
        df = pd.DataFrame(courses)
        
        summary = df.groupby('semester').agg({
            'course_code': 'count',
            'credits': 'sum'
        }).reset_index()
        
        summary.columns = ['Semester', 'Number of Courses', 'Total Credits']
        
        # Add totals row
        totals = pd.DataFrame([{
            'Semester': 'TOTAL',
            'Number of Courses': summary['Number of Courses'].sum(),
            'Total Credits': summary['Total Credits'].sum()
        }])
        
        summary = pd.concat([summary, totals], ignore_index=True)
        
        return summary
    
    def _create_program_summary(self, structured_data: Dict) -> pd.DataFrame:
        """Create program-level summary."""
        summary_data = {
            'Metric': [
                'Total Courses',
                'Total Credits',
                'Average Credits per Semester',
                'Number of Semesters',
                'Generated Date'
            ],
            'Value': [
                str(structured_data.get('total_courses', 0)),
                str(structured_data.get('total_credits', 0)),
                str(round(structured_data.get('total_credits', 0) / max(len(set(c['semester'] for c in structured_data['courses'])), 1), 1)),
                str(len(set(c['semester'] for c in structured_data['courses']))),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        return pd.DataFrame(summary_data)
    
    def _create_metadata(self, metadata: Optional[Dict]) -> pd.DataFrame:
        """Create metadata sheet."""
        if not metadata:
            metadata = {}
        
        meta_data = {
            'Field': [
                'Generated Date',
                'Tool Version',
                'DegreeWorks File',
                'Study Plan File',
                'Schedule File'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                metadata.get('version', 'Smart Class Planner v1.0'),
                metadata.get('degreeworks_file', 'N/A'),
                metadata.get('study_plan_file', 'N/A'),
                metadata.get('schedule_file', 'N/A')
            ]
        }
        
        return pd.DataFrame(meta_data)
    
    def _apply_formatting(self, workbook: openpyxl.Workbook):
        """Apply professional formatting to workbook."""
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Format headers
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')


# Utility function for quick export
def quick_export(plan_data: Any, file_path: str, metadata: Optional[Dict] = None):
    """
    Quick export function for simple use cases.
    
    Args:
        plan_data: Plan data to export
        file_path: Output file path
        metadata: Optional metadata
    """
    exporter = ExcelExporter()
    exporter.export_plan(plan_data, file_path, metadata)


if __name__ == "__main__":
    # Example usage
    print("Excel Exporter Module - Example Usage")
    print("=" * 50)
    
    # Example: Export sample course data
    sample_courses = [
        {
            'semester': 'Fall 2025',
            'course_code': 'CPSC 6177',
            'course_name': 'Software Engineering',
            'credits': 3,
            'prerequisites': 'None',
            'status': 'Planned'
        },
        {
            'semester': 'Fall 2025',
            'course_code': 'CPSC 6000',
            'course_name': 'Algorithms',
            'credits': 3,
            'prerequisites': 'None',
            'status': 'Planned'
        },
        {
            'semester': 'Spring 2026',
            'course_code': 'CPSC 6178',
            'course_name': 'Advanced Software Engineering',
            'credits': 3,
            'prerequisites': 'CPSC 6177',
            'status': 'Planned'
        }
    ]
    
    exporter = ExcelExporter()
    exporter.export_simple(sample_courses, 'sample_plan.xlsx')
    print("Sample plan exported to 'sample_plan.xlsx'")